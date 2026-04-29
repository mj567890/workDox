from io import BytesIO


def extract_text_from_docx(data: bytes) -> str:
    try:
        from docx import Document as DocxDocument
        doc = DocxDocument(BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception:
        return ""


def extract_text_from_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(data), read_only=True)
        texts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except Exception:
        return ""


def extract_text_from_pdf(data: bytes) -> str:
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(BytesIO(data)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return "\n".join(text_parts)
    except Exception:
        return ""


def extract_text_from_txt(data: bytes) -> str:
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return data.decode("gbk")
        except Exception:
            return ""


def extract_text(file_type: str, data: bytes) -> str:
    extractors = {
        "docx": extract_text_from_docx,
        "doc": extract_text_from_docx,
        "xlsx": extract_text_from_xlsx,
        "xls": extract_text_from_xlsx,
        "pdf": extract_text_from_pdf,
        "txt": extract_text_from_txt,
        "md": extract_text_from_txt,
        "csv": extract_text_from_txt,
    }
    extractor = extractors.get(file_type)
    if extractor:
        return extractor(data)
    return ""
